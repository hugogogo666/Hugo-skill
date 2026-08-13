#!/usr/bin/env python3
"""Build a formatted Amazon ASIN advertising plan workbook from JSON."""
import argparse, json, os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

NAVY="17365D"; BLUE="1F4E78"; ORANGE="ED7D31"; GREEN="70AD47"
WHITE="FFFFFF"; LIGHT_BLUE="D9EAF7"; YELLOW="FFF2CC"; LIGHT_GREEN="E2F0D9"; LIGHT_RED="F4CCCC"
THIN=Side(style="thin",color="C9C9C9")


def title(ws, text, subtitle="", freeze="A4"):
    ws.sheet_view.showGridLines=False; ws.freeze_panes=freeze
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=10)
    c=ws.cell(1,1,text); c.font=Font(name="Microsoft YaHei",size=16,bold=True,color=WHITE)
    c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(vertical="center")
    ws.row_dimensions[1].height=30
    if subtitle:
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=10)
        c=ws.cell(2,1,subtitle); c.font=Font(name="Microsoft YaHei",size=10,color="666666",italic=True)
        c.alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[2].height=28


def header(ws,row,labels,color=BLUE):
    for i,label in enumerate(labels,1):
        c=ws.cell(row,i,label); c.font=Font(name="Microsoft YaHei",bold=True,color=WHITE)
        c.fill=PatternFill("solid",fgColor=color); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
    ws.row_dimensions[row].height=28


def body(ws,start,end,cols):
    for row in ws.iter_rows(min_row=start,max_row=end,min_col=1,max_col=cols):
        for c in row:
            c.font=Font(name="Microsoft YaHei",size=10)
            c.alignment=Alignment(vertical="top",wrap_text=True)
            c.border=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
            if c.row%2==0: c.fill=PatternFill("solid",fgColor="F8FBFD")
        ws.row_dimensions[row[0].row].height=38


def widths(ws, values):
    for col,w in values.items(): ws.column_dimensions[col].width=w


def add_rows(ws, rows, keys, start=5):
    for r_idx,item in enumerate(rows,start):
        for c_idx,key in enumerate(keys,1): ws.cell(r_idx,c_idx,item.get(key,""))
    return start+len(rows)-1


def prepare(ws):
    ws.sheet_properties.pageSetUpPr.fitToPage=True; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.page_setup.orientation="landscape"; ws.page_margins.left=.25; ws.page_margins.right=.25


def load(path):
    with open(path,"r",encoding="utf-8") as f: return json.load(f)


def build(d, out):
    asin=d.get("asin","ASIN"); market=d.get("marketplace","US"); snap=d.get("snapshot_date",datetime.now().date().isoformat())
    subtitle=f"站点：Amazon {market}｜数据快照：{snap}｜最终竞价和预算以卖家后台表现及真实利润模型为准。"
    wb=Workbook(); wb.remove(wb.active)

    # Overview
    ws=wb.create_sheet("01_ASIN现状诊断"); title(ws,f"{asin}｜亚马逊广告推进方案",subtitle)
    labs=["维度","当前数据/状态","经营判断","对广告计划的影响"]; header(ws,4,labs)
    end=add_rows(ws,d.get("overview",[]),["dimension","value","judgment","impact"])
    if end>=5: body(ws,5,end,4); ws.auto_filter.ref=f"A4:D{end}"
    widths(ws,{"A":22,"B":38,"C":40,"D":55})
    if d.get("product_url") and end>=5:
        for r in range(5,end+1):
            if str(ws.cell(r,2).value)==asin: ws.cell(r,2).hyperlink=d["product_url"]; ws.cell(r,2).style="Hyperlink"
    ws.merge_cells(start_row=max(end+3,8),start_column=1,end_row=max(end+3,8),end_column=4)
    c=ws.cell(max(end+3,8),1,"核心结论："+d.get("summary","请结合数据建立核心词、效率词、探索词和商品定向四类广告。"))
    c.fill=PatternFill("solid",fgColor="FCE4D6"); c.font=Font(name="Microsoft YaHei",bold=True,size=11); c.alignment=Alignment(wrap_text=True)

    # Keywords
    ws=wb.create_sheet("02_关键词分层与竞价"); title(ws,"关键词分层、匹配方式与初始竞价",subtitle)
    labs=["优先级","关键词","中文意图","广告角色","匹配方式","建议初始竞价($)","数据依据","投放动作","风险提示"]; header(ws,4,labs)
    end=add_rows(ws,d.get("keywords",[]),["priority","keyword","intent","role","match","bid","evidence","action","risk"])
    if end>=5: body(ws,5,end,9); ws.auto_filter.ref=f"A4:I{end}"
    widths(ws,{"A":10,"B":34,"C":25,"D":18,"E":14,"F":18,"G":42,"H":42,"I":38})

    # Campaigns
    ws=wb.create_sheet("03_广告活动架构"); title(ws,"建议广告活动架构",subtitle)
    labs=["序号","活动命名建议","广告类型","主要目标","主要定向/关键词","竞价策略","初始竞价($)","预算占比","$60日预算","管理规则"]; header(ws,4,labs)
    camps=d.get("campaigns",[])
    for idx,item in enumerate(camps,5):
        vals=[idx-4,item.get("name",""),item.get("type",""),item.get("goal",""),item.get("targets",""),item.get("strategy",""),item.get("bid",""),item.get("budget_share",0),f"=H{idx}*60",item.get("rule","")]
        for col,val in enumerate(vals,1): ws.cell(idx,col,val)
        ws.cell(idx,8).number_format="0%"; ws.cell(idx,9).number_format='$0.00'
    end=4+len(camps)
    if end>=5: body(ws,5,end,10); ws.auto_filter.ref=f"A4:J{end}"
    widths(ws,{"A":8,"B":28,"C":18,"D":18,"E":50,"F":22,"G":17,"H":13,"I":14,"J":45})

    # Economics
    eco=d.get("economics",{})
    ws=wb.create_sheet("04_预算与盈亏模型"); title(ws,"预算、目标CPA与最高可承受CPC",subtitle)
    header(ws,4,["参数","输入值","说明"])
    price=eco.get("price",19.99); acos=eco.get("target_acos",.30); cvr=eco.get("expected_cvr",.15); orders=eco.get("target_orders_day",5)
    rows=[("实际成交价($)",price,"按实际Buy Box成交价计算"),("目标ACoS",acos,"结合成本和目标利润修正"),("预期广告CVR",cvr,"用后台真实转化率替换"),("目标CPA($)","=B5*B6","单个广告订单可承受花费"),("最高可承受CPC($)","=B5*B6*B7","长期CPC上限"),("目标广告订单/天",orders,"可编辑"),("理论日预算($)","=B8*B9","目标CPA×目标订单")]
    for r_idx,row in enumerate(rows,5):
        for c_idx,val in enumerate(row,1): ws.cell(r_idx,c_idx,val)
    body(ws,5,11,3); widths(ws,{"A":30,"B":20,"C":70})
    for cell in ["B5","B6","B7","B9"]: ws[cell].fill=PatternFill("solid",fgColor=YELLOW)
    ws["B6"].number_format="0%"; ws["B7"].number_format="0%"
    for cell in ["B5","B8","B10","B11"]: ws[cell].number_format='$0.00'
    header(ws,14,["实际CPC($)","最多点击数","所需CVR","判断"],ORANGE)
    for r,cpc in enumerate([.55,.70,.85,1.07,1.25],15):
        ws.cell(r,1,cpc); ws.cell(r,2,f"=$B$8/A{r}"); ws.cell(r,3,f"=A{r}/($B$5*$B$6)"); ws.cell(r,4,f'=IF(C{r}<=$B$7,"可承受","需提高CVR或降低CPC")')
        ws.cell(r,1).number_format='$0.00'; ws.cell(r,2).number_format='0.0'; ws.cell(r,3).number_format='0.0%'
    body(ws,15,19,4); ws.conditional_formatting.add("C15:C19",ColorScaleRule(start_type='min',start_color=LIGHT_GREEN,mid_type='percentile',mid_value=50,mid_color=YELLOW,end_type='max',end_color=LIGHT_RED))

    # Roadmap
    ws=wb.create_sheet("05_阶段推进计划"); title(ws,"广告阶段推进路线图",subtitle)
    labs=["阶段","时间","核心目标","具体动作","关键产出","验收标准","禁止事项"]; header(ws,4,labs)
    end=add_rows(ws,d.get("roadmap",[]),["stage","time","goal","actions","deliverable","acceptance","avoid"])
    if end>=5: body(ws,5,end,7); ws.auto_filter.ref=f"A4:G{end}"
    widths(ws,{"A":14,"B":14,"C":25,"D":65,"E":36,"F":48,"G":48})
    rt=d.get("ranking_targets",[]); start=max(end+3,10); header(ws,start,["关键词","当前自然位置信号","阶段目标","优先级","备注"],GREEN)
    for idx,item in enumerate(rt,start+1):
        for col,key in enumerate(["keyword","current","target","priority","note"],1): ws.cell(idx,col,item.get(key,""))
    if rt: body(ws,start+1,start+len(rt),5)

    # Rules
    ws=wb.create_sheet("06_优化与止损规则"); title(ws,"广告优化决策规则",subtitle)
    labs=["场景","识别条件","优先判断","执行动作","调整幅度/止损线","复查周期"]; header(ws,4,labs)
    end=add_rows(ws,d.get("rules",[]),["scenario","condition","check","action","range","review"])
    if end>=5: body(ws,5,end,6); ws.auto_filter.ref=f"A4:F{end}"
    widths(ws,{"A":25,"B":38,"C":38,"D":48,"E":35,"F":15})

    # Negatives
    ws=wb.create_sheet("07_否定词与隔离词"); title(ws,"否定关键词与独立测试词",subtitle)
    labs=["关键词/词根","类型","否定匹配方式","处理建议","原因","适用活动"]; header(ws,4,labs)
    end=add_rows(ws,d.get("negatives",[]),["term","type","match","action","reason","scope"])
    if end>=5: body(ws,5,end,6); ws.auto_filter.ref=f"A4:F{end}"
    widths(ws,{"A":38,"B":22,"C":22,"D":24,"E":55,"F":32})

    # Listing
    ws=wb.create_sheet("08_Listing承接清单"); title(ws,"广告投放前的Listing承接清单",subtitle)
    labs=["页面位置","承接需求","建议表达","检查状态","优先级","备注"]; header(ws,4,labs)
    end=add_rows(ws,d.get("listing",[]),["position","demand","suggestion","status","priority","note"])
    if end>=5: body(ws,5,end,6); ws.auto_filter.ref=f"A4:F{end}"
    widths(ws,{"A":18,"B":32,"C":75,"D":22,"E":12,"F":45})
    if end>=5:
        dv=DataValidation(type="list",formula1='"待检查,待强化,已完成,不适用"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"D5:D{end}")

    # Monitoring
    ws=wb.create_sheet("09_广告监测模板"); title(ws,"广告监测与操作记录模板",subtitle)
    labs=["日期","活动","广告组","关键词/ASIN","匹配方式","活动目的","出价($)","日预算($)","曝光","点击","CTR","CPC($)","花费($)","订单","CVR","广告销售额($)","ACoS","ROAS","自然排名","价格($)","Coupon","操作记录","结论","复查日期"]; header(ws,4,labs)
    for row in range(5,105):
        ws.cell(row,11,f'=IFERROR(J{row}/I{row},"")'); ws.cell(row,12,f'=IFERROR(M{row}/J{row},"")')
        ws.cell(row,15,f'=IFERROR(N{row}/J{row},"")'); ws.cell(row,17,f'=IFERROR(M{row}/P{row},"")'); ws.cell(row,18,f'=IFERROR(P{row}/M{row},"")')
        for col in [7,8,12,13,16,20]: ws.cell(row,col).number_format='$0.00'
        for col in [11,15,17]: ws.cell(row,col).number_format='0.0%'
    body(ws,5,104,24); ws.auto_filter.ref="A4:X104"
    widths(ws,{"A":13,"B":27,"C":18,"D":34,"E":13,"F":13,"G":12,"H":13,"I":11,"J":10,"K":11,"L":11,"M":12,"N":10,"O":11,"P":15,"Q":11,"R":10,"S":12,"T":11,"U":12,"V":38,"W":15,"X":15})
    for rng,formula in [("E5:E104",'"精准,词组,广泛,自动紧密,自动宽泛,自动替代,商品定向"'),("F5:F104",'"盈利,排名,探索,增长,竞品"'),("W5:W104",'"待观察,加价,降价,加预算,收词,否定,暂停,维持"')]:
        dv=DataValidation(type="list",formula1=formula,allow_blank=True); ws.add_data_validation(dv); dv.add(rng)

    # KPI
    ws=wb.create_sheet("10_KPI与数据缺口"); title(ws,"阶段KPI、数据缺口与后续精算项",subtitle)
    labs=["类别","指标/数据","当前值/目标","用途","优先级","获取方式/说明"]; header(ws,4,labs)
    end=add_rows(ws,d.get("kpis",[]),["category","metric","current_target","purpose","priority","source"])
    if end>=5: body(ws,5,end,6); ws.auto_filter.ref=f"A4:F{end}"
    widths(ws,{"A":18,"B":36,"C":38,"D":38,"E":14,"F":58})

    for ws in wb.worksheets: prepare(ws)
    wb.properties.title=f"{asin} Amazon广告推进方案"; wb.properties.creator="Wenmai AI"
    os.makedirs(os.path.dirname(os.path.abspath(out)),exist_ok=True); wb.save(out)
    check=load_workbook(out,data_only=False)
    if len(check.sheetnames)!=10 or check["04_预算与盈亏模型"]["B8"].value!="=B5*B6": raise RuntimeError("workbook validation failed")


def main():
    p=argparse.ArgumentParser(); p.add_argument("--input",required=True); p.add_argument("--output",required=True); a=p.parse_args()
    build(load(a.input),a.output); print(os.path.abspath(a.output))
if __name__=="__main__": main()
