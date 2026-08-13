#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Validate the two-stage Ozon -> 1688 workflow gate.

It intentionally does not call any external service. The matching stage requires
an explicit --confirm flag and a completed selection sheet.
"""
import argparse, json
from pathlib import Path
from openpyxl import load_workbook

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--input', required=True)
    ap.add_argument('--stage', choices=['check','matching'], default='check')
    ap.add_argument('--confirm', action='store_true')
    args=ap.parse_args()
    path=Path(args.input)
    if not path.exists():
        raise SystemExit('SCREENING_REQUIRED: 阶段一候选文件不存在')
    wb=load_workbook(path, read_only=True, data_only=True)
    required = '1688搜图优先30款' if '1688搜图优先30款' in wb.sheetnames else ('A级优先商品' if 'A级优先商品' in wb.sheetnames else None)
    if not required:
        raise SystemExit('SCREENING_REQUIRED: 未找到阶段一候选商品工作表')
    ws=wb[required]
    count=max(0, ws.max_row-1)
    if count == 0:
        raise SystemExit('SCREENING_REQUIRED: 阶段一候选商品为空')
    if args.stage == 'matching' and not args.confirm:
        raise SystemExit('SELECTION_CONFIRMATION_REQUIRED: 必须先获得用户明确确认，再进入1688搜图')
    result={'status':'SELECTION_CONFIRMED' if args.stage=='matching' else 'SCREENING_COMPLETED','sheet':required,'selected_count':count}
    print(json.dumps(result,ensure_ascii=False,indent=2))
if __name__=='__main__': main()
